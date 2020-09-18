'''Васю назначили завхозом в туристической группе и он подошёл к подготовке ответственно,
составив справочник продуктов с указанием калорийности на 100 грамм, а также содержание белков,
жиров и углеводов на 100 грамм продукта. Ему не удалось найти всю информацию, поэтому некоторые
ячейки остались незаполненными (можно считать их значение равным нулю). Также он использовал к
акой-то странный офисный пакет и разделял целую и дробную часть чисел запятой. Таблица доступа
по ссылке https://stepik.org/media/attachments/lesson/245290/trekking1.xlsx

Вася хочет минимизировать вес продуктов и для этого брать самые калорийные продукты.
Помогите ему и упорядочите продукты по убыванию калорийности. В случае, если продукты имеют
одинаковую калорийность - упорядочите их по названию. В качестве ответа необходимо сдать названия
продуктов, по одному в строке. '''

import openpyxl #импортируем модуль работы с xml файлами
def key_name(x):
    return x[1]

def key_cal(x):
    return x[0]
fin = openpyxl.load_workbook("trekking1.xlsx")
fout = open("task_15_output.txt", "w", encoding="utf8")

sheet = fin["Справочник"]

callories=[]

for rw in range(2, sheet.max_row+1):
    calorii = float(sheet.cell(row=rw, column=2).value)
    product = sheet.cell(row=rw, column=1).value
    callories.append((calorii, product))
print(callories)
callories.sort(key=key_name)
print(callories)
callories.sort(reverse=True, key=key_cal)
#sorted_cal = sorted(callories, key=key_name)
print(callories)

for prod in callories:
    print(prod[1], file=fout)

fout.close()

