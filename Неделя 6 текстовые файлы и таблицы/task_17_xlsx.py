'''Васю назначили завхозом в туристической группе и он подошёл к подготовке ответственно,
составив справочник продуктов с указанием калорийности на 100 грамм, а также содержание белков,
жиров и углеводов на 100 грамм продукта. Ему не удалось найти всю информацию, поэтому некоторые
ячейки остались незаполненными (можно считать их значение равным нулю). Также он использовал какой-то
странный офисный пакет и разделял целую и дробную часть чисел запятой. Таблица доступна
по ссылке https://stepik.org/media/attachments/lesson/258931/trekking3.xlsx

Вася составил раскладку по продуктам на весь поход (она на листе "Раскладка") с указанием номера дня,
названия продукта и его количества в граммах. Для каждого дня посчитайте 4 числа: суммарную калорийность и
граммы белков, жиров и углеводов.
Числа округлите до целых вниз и введите через пробел. Информация о каждом дне должна выводиться в отдельной строке. '''

import openpyxl #импортируем модуль работы с xml файлами

callories = 0
belok = 0
zhir = 0
ugl = 0

wb = openpyxl.load_workbook("trekking3.xlsx")
fout = open("task_17_output.txt", "w", encoding="utf8")

sheet = wb["Справочник"]

products={}
for rw in range(2, sheet.max_row+1):
    temp=[] #нулевой список по калорийности, жирам, углеводам
    for col in range(2, sheet.max_column+1):
        if not sheet.cell(row=rw, column = col).value:
            temp.append(0)
        else:
            temp.append(sheet.cell(row=rw, column = col).value)
    products[sheet.cell(row=rw, column=1).value] = temp

sheet = wb["Раскладка"]

days={}
for rw in range(2, sheet.max_row+1): # пробегаем по каждой строчке
    if sheet.cell(row=rw, column=1).value not in days: #проверяем на наличие дня в днях
        days[sheet.cell(row=rw, column=1).value] = [0, 0, 0, 0] #если нет - создаем день с нулевыми значениями
    for i in range(0, 4): #в любом случае, необходимо добавить каллории
        days[sheet.cell(row=rw, column=1).value][i] += products[sheet.cell(row=rw, column=2).value][i] * sheet.cell(row=rw, column=3).value / 100


for i in days:
    print(*list(map(int, days[i])), file = fout)

fout.close()

