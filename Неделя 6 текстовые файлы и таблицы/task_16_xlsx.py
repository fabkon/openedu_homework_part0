'''Васю назначили завхозом в туристической группе и он подошёл к подготовке ответственно,
составив справочник продуктов с указанием калорийности на 100 грамм, а также содержание белков,
жиров и углеводов на 100 грамм продукта. Ему не удалось найти всю информацию, поэтому некоторые
 ячейки остались незаполненными (можно считать их значение равным нулю). Также он использовал
 какой-то странный офисный пакет и разделял целую и дробную часть чисел запятой. Таблица доступна
 по ссылке https://stepik.org/media/attachments/lesson/245290/trekking2.xlsx

Вася составил раскладку по продуктам на один день (она на листе "Раскладка") с указанием названия продукта
и его количества в граммах. Посчитайте 4 числа:
суммарную калорийность и граммы белков, жиров и углеводов. Числа округлите до целых вниз и введите через пробел. '''

import openpyxl #импортируем модуль работы с xml файлами

callories = 0
belok = 0
zhir = 0
ugl = 0

wb = openpyxl.load_workbook("trekking2.xlsx")
fout = open("task_16_output.txt", "w", encoding="utf8")

sheet = wb["Справочник"]

products={}
for rw in range(2, sheet.max_row+1):
    temp=[]
    for col in range(2, sheet.max_column+1):
        if not sheet.cell(row=rw, column = col).value:
            temp.append(0)
        else:
            temp.append(sheet.cell(row=rw, column = col).value)

    products[sheet.cell(row=rw, column=1).value] = temp
#print(products)

sheet = wb["Раскладка"]
summurize = []
for rw in range(2, sheet.max_row+1):
    callories += products[sheet.cell(row=rw, column=1).value][0] * sheet.cell(row=rw, column=2).value / 100
    belok += products[sheet.cell(row=rw, column=1).value][1] * sheet.cell(row=rw, column=2).value / 100
    zhir += products[sheet.cell(row=rw, column=1).value][2] * sheet.cell(row=rw, column=2).value / 100
    ugl += products[sheet.cell(row=rw, column=1).value][3] * sheet.cell(row=rw, column=2).value / 100

print(*list(map(int,(callories, belok, zhir, ugl))), sep=' ', file=fout)
fout.close()

