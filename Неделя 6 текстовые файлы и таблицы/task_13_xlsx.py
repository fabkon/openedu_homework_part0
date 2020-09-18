import openpyxl #импортируем модуль работы с xml файлами
import requests #импортируем модуль загрузки файла из интернета
import zipfile #импортируем модуль работы с архивами

#help(zipfile)
"""
#СОХРАНЕНИЕ ФАЙЛА ИЗ ИНТЕРНЕТА
f=open(r'rogaikopyta_test1.zip',"wb") #открываем файл для записи, в режиме wb
ufr = requests.get("https://stepik.org/media/attachments/lesson/245299/rogaikopyta.zip") #делаем запрос
f.write(ufr.content) #записываем содержимое в файл; как видите - content запроса
f.close()

#РАЗАРХИВИРОВАНИЕ ФАЙЛА в отдельную папку
z = zipfile.ZipFile('rogaikopyta_test1.zip', 'r')
z.extractall(path="test_2")
z.close()
"""

#выгрузка данных из таблиц в файл
vedomost = []
fin = open("vedomost.txt", "w", encoding="utf8")

for i in range(1000):
    name_of_file = 'test_2/'+str(i+1)+'.xlsx'
    wb = openpyxl.load_workbook(name_of_file)
    sh = wb.active
    name = sh.cell(row=2, column=2).value
    salary = sh.cell(row=2, column=4).value
    vedomost.append((name,salary))

vedomost.sort()
for i in vedomost:
    print(*i, sep=" ", file = fin)
fin.close()
