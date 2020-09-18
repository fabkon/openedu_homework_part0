'''Вася планирует карьеру и переезд. Для это составил таблицу, в которой для каждого региона записал зарплаты
для разных интересные ему профессий. Таблица доступна по ссылке
 https://stepik.org/media/attachments/lesson/245267/salaries.xlsx. Выведите название региона с самой
 высокой медианной зарплатой (медианой называется элемент, стоящий в середине массива после его упорядочивания)
и, через пробел, название профессии с самой высокой средней зарплатой по всем регионам. '''

import openpyxl #импортируем модуль работы с xml файлами
import requests #импортируем модуль загрузки файла из интернета

"""
#СОХРАНЕНИЕ ФАЙЛА ИЗ ИНТЕРНЕТА
f = open('task_14.xlsx',"wb") #открываем файл для записи, в режиме wb (запись в двоичном виде)
ufr = requests.get("https://stepik.org/media/attachments/lesson/245267/salaries.xlsx") #делаем запрос
f.write(ufr.content) #записываем содержимое в файл; как видите - content запроса
f.close()
"""

fin = openpyxl.load_workbook("task_14.xlsx")
fout = open("task_14_output.txt", "w", encoding="utf8")
#sheet = fin.active
sheet = fin["Лист1"]

cities={}

for i in range(2, sheet.max_row+1):
    city = sheet.cell(row=i, column=1).value
    cities[city]=[]
    for j in range(2, sheet.max_column+1):
        cities[city].append(sheet.cell(row=i, column=j).value)
    cities[city].sort()
#print(cities)

medians=[]
for city in cities:
    medians.append((cities[city][len(cities[city])//2], city))
print(max(medians)[1], end=' ', file=fout)

proffs={}

for j in range(2, sheet.max_column+1):
    proff = sheet.cell(row=1, column=j).value
    proffs[proff]=[]
    for i in range(2, sheet.max_row+1):
        proffs[proff].append(sheet.cell(row=i, column=j).value)

#print(proffs)

average_sal=[]
for proff in proffs:
    average_sal.append((sum(proffs[proff]), proff))
print(max(average_sal)[1], file=fout)


fout.close()

"""
#выгрузка данных из таблиц в файл
vedomost = []
fin = open("vedomost.txt", "w", encoding="utf8")

for i in range(1000):
    name_of_file = 'test_2/'+str(i+1)+'.xlsx'
    wb = openpyxl.load_workbook(name_of_file)
    sh = wb.active
    name = sh.cell(row=2, column=2).value
    salary = sh.cell(row=2, column=4).value
    vedomost.append((name,salary))

vedomost.sort()
for i in vedomost:
    print(*i, sep=" ", file = fin)
fin.close()
"""